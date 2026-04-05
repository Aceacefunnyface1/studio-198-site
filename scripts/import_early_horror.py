from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("/Users/bradbehnke/Desktop/movie_posters/1689-1968.xlsx")
FIX_WORKBOOK_PATH = Path("/Users/bradbehnke/Downloads/poster_fix_70.xlsx")
SOURCE_POSTER_DIR = Path("/Users/bradbehnke/Desktop/movie_posters/posters")
DATA_PATH = Path(__file__).resolve().parents[1] / "data" / "site-data.json"
DEST_POSTER_DIR = Path(__file__).resolve().parents[1] / "public" / "posters" / "early-horror"
REPORT_PATH = Path(__file__).resolve().parents[1] / "reports" / "early-horror-import-report.md"

COLLECTION = "early-horror"
START_YEAR = 1896
END_YEAR = 1969
PARENTHESES_PATTERN = re.compile(r"\([^()]*\)")
BRACKET_PATTERN = re.compile(r"\[[^\]]*\]")
AKA_MARKERS = (
    "aka",
    "a.k.a.",
    "usa title:",
    "us title:",
    "u.s. title:",
    "alt title:",
)


def normalize_unicode(value: object) -> str:
    text = str(value or "").strip()
    text = text.replace("’", "'").replace("‘", "'").replace("“", '"').replace("”", '"')
    text = text.replace("–", "-").replace("—", "-")
    normalized = unicodedata.normalize("NFKD", text)
    return normalized.encode("ascii", "ignore").decode("ascii")


def normalize_key(value: object) -> str:
    text = normalize_unicode(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def slugify(value: object) -> str:
    text = normalize_unicode(value).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-{2,}", "-", text).strip("-")


def collapse_spaces(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def strip_parenthetical_sections(value: object) -> str:
    text = normalize_unicode(value)
    previous = None
    while previous != text:
      previous = text
      text = PARENTHESES_PATTERN.sub(" ", text)
    text = BRACKET_PATTERN.sub(" ", text)
    return collapse_spaces(text)


def extract_aliases_from_parentheticals(value: object) -> list[str]:
    text = normalize_unicode(value)
    aliases: list[str] = []

    for match in PARENTHESES_PATTERN.findall(text):
        inner = match[1:-1].strip()
        inner_lower = inner.lower()
        alias = inner

        for marker in AKA_MARKERS:
            if marker in inner_lower:
                start = inner_lower.index(marker) + len(marker)
                alias = inner[start:].strip(" :-")
                break

        cleaned = collapse_spaces(alias)
        if cleaned:
            aliases.append(cleaned)

    return aliases


def split_tags(value: object) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item and item.strip()]


def parse_rating(value: object) -> float:
    try:
        rating = float(str(value).strip())
    except (TypeError, ValueError):
        return 3.0

    return max(0.0, min(5.0, rating))


def parse_year(value: object) -> int | None:
    try:
        year = int(str(value).strip())
    except (TypeError, ValueError):
        return None

    if START_YEAR <= year <= END_YEAR:
        return year

    return year


def verdict_for_rating(rating: float) -> str:
    if rating >= 4.0:
        return "🔥"
    if rating >= 3.0:
        return "👀"
    if rating >= 1.1:
        return "❌"
    return "💩"


def normalize_reviewer(value: object) -> str:
    text = collapse_spaces(value)
    upper = text.upper()

    if "MINDY" in upper or "MANDY" in upper:
        return "Mindy"
    if "LEE" in upper:
        return "Leeanne"
    return "Ace"


def strip_trailing_year_slug(slug: str) -> str:
    return re.sub(r"-(18|19|20)\d{2}$", "", slug)


def candidate_keys_from_text(value: object) -> set[str]:
    candidates = set()
    raw = collapse_spaces(value)

    if raw:
        candidates.add(normalize_key(raw))
        stripped = strip_parenthetical_sections(raw)
        if stripped:
            candidates.add(normalize_key(stripped))
        for alias in extract_aliases_from_parentheticals(raw):
            candidates.add(normalize_key(alias))

    return {candidate for candidate in candidates if candidate}


def candidate_keys_from_slug(value: object) -> set[str]:
    slug = slugify(value)
    if not slug:
        return set()

    candidates = {
        normalize_key(slug),
        normalize_key(strip_trailing_year_slug(slug)),
    }
    return {candidate for candidate in candidates if candidate}


def build_poster_lookup() -> dict[str, Path]:
    lookup: dict[str, Path] = {}

    for poster_path in SOURCE_POSTER_DIR.iterdir():
        if not poster_path.is_file():
            continue

        stem = poster_path.stem
        for key in candidate_keys_from_text(stem) | candidate_keys_from_slug(stem):
            lookup.setdefault(key, poster_path)

    return lookup


def read_main_rows() -> list[dict[str, object]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, object]] = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values:
            continue

        row = dict(zip(headers, values))
        if not row.get("Movie Title") or not row.get("Slug"):
            continue
        rows.append(row)

    return rows


def read_fix_rows() -> tuple[dict[str, dict[str, object]], dict[str, dict[str, object]]]:
    if not FIX_WORKBOOK_PATH.exists():
        return {}, {}

    workbook = load_workbook(FIX_WORKBOOK_PATH, read_only=True)
    sheet = workbook["Poster_Fix_70"]
    headers = [cell for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    by_title: dict[str, dict[str, object]] = {}
    by_slug: dict[str, dict[str, object]] = {}

    for values in sheet.iter_rows(min_row=3, values_only=True):
        if not values or not values[0]:
            continue

        row = dict(zip(headers, values))
        original_title_key = normalize_key(row.get("Original Title"))
        existing_slug = slugify(row.get("Existing DB Slug"))

        if original_title_key:
            by_title[original_title_key] = row
        if existing_slug:
            by_slug[existing_slug] = row

    return by_title, by_slug


def find_fix_row(
    row: dict[str, object],
    fixes_by_title: dict[str, dict[str, object]],
    fixes_by_slug: dict[str, dict[str, object]],
) -> dict[str, object] | None:
    title_key = normalize_key(row.get("Movie Title"))
    slug_key = slugify(row.get("Slug"))

    return fixes_by_slug.get(slug_key) or fixes_by_title.get(title_key)


def choose_display_title(source_row: dict[str, object], fix_row: dict[str, object] | None) -> str:
    title = fix_row.get("Fixed Title") if fix_row else None
    title = collapse_spaces(title or source_row.get("Movie Title"))
    return title or "Untitled"


def choose_slug(source_row: dict[str, object], fix_row: dict[str, object] | None) -> str:
    value = fix_row.get("Existing DB Slug") if fix_row else None
    slug = slugify(value or source_row.get("Slug") or choose_display_title(source_row, fix_row))
    return slug


def find_matching_poster(
    source_row: dict[str, object],
    fix_row: dict[str, object] | None,
    poster_lookup: dict[str, Path],
) -> Path | None:
    candidates: list[str] = []
    display_title = choose_display_title(source_row, fix_row)
    slug = choose_slug(source_row, fix_row)

    text_candidates = [
        source_row.get("Movie Title"),
        display_title,
        fix_row.get("Original Title") if fix_row else None,
        fix_row.get("Fixed Title") if fix_row else None,
        fix_row.get("Alt Title 1") if fix_row else None,
        fix_row.get("Alt Title 2") if fix_row else None,
    ]
    slug_candidates = [
        source_row.get("Slug"),
        slug,
        fix_row.get("Poster Match Slug") if fix_row else None,
        fix_row.get("Alt Slug 1") if fix_row else None,
        fix_row.get("Alt Slug 2") if fix_row else None,
    ]

    for value in text_candidates:
        candidates.extend(candidate_keys_from_text(value))
    for value in slug_candidates:
        candidates.extend(candidate_keys_from_slug(value))

    for candidate in candidates:
        if candidate and candidate in poster_lookup:
            return poster_lookup[candidate]

    return None


def choose_row_value(
    source_row: dict[str, object],
    fix_row: dict[str, object] | None,
    key: str,
) -> object:
    if fix_row and fix_row.get(key) not in (None, ""):
        return fix_row.get(key)
    return source_row.get(key)


def clear_destination_directory() -> None:
    DEST_POSTER_DIR.mkdir(parents=True, exist_ok=True)

    for existing_path in DEST_POSTER_DIR.iterdir():
        if existing_path.is_file():
            existing_path.unlink()


def build_report(
    imported_count: int,
    live_visible_count: int,
    hidden_pending_count: int,
    missing_titles: list[str],
) -> None:
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Early Horror Import Report",
        "",
        f"- Imported count: {imported_count}",
        f"- Live visible count: {live_visible_count}",
        f"- Hidden pending-poster count: {hidden_pending_count}",
        f"- Titles still failing after normalization/alias pass: {len(missing_titles)}",
        "",
        "## Still Missing Posters",
        "",
    ]

    if missing_titles:
        lines.extend(f"- {title}" for title in missing_titles)
    else:
        lines.append("- None")

    REPORT_PATH.write_text("\n".join(lines) + "\n")


def main() -> None:
    main_rows = read_main_rows()
    fixes_by_title, fixes_by_slug = read_fix_rows()
    poster_lookup = build_poster_lookup()
    site_data = json.loads(DATA_PATH.read_text())
    reviews = site_data["reviews"]
    reviews_by_slug = {review["slug"]: review for review in reviews}
    now = datetime.now(timezone.utc).isoformat()

    clear_destination_directory()

    imported_count = 0
    live_visible_count = 0
    hidden_pending_count = 0
    missing_titles: list[str] = []
    ordered_imported_slugs: list[str] = []

    for source_row in main_rows:
        fix_row = find_fix_row(source_row, fixes_by_title, fixes_by_slug)
        title = choose_display_title(source_row, fix_row)
        slug = choose_slug(source_row, fix_row)
        year = parse_year(choose_row_value(source_row, fix_row, "Year"))
        rating = parse_rating(choose_row_value(source_row, fix_row, "Rating"))
        poster_path = find_matching_poster(source_row, fix_row, poster_lookup)
        existing_review = reviews_by_slug.get(slug)
        imported_count += 1

        poster_image = ""
        if poster_path is not None:
            destination_name = f"{slug}{poster_path.suffix.lower()}"
            destination_path = DEST_POSTER_DIR / destination_name
            shutil.copy2(poster_path, destination_path)
            poster_image = f"/posters/early-horror/{destination_name}"
            live_visible_count += 1
        else:
            hidden_pending_count += 1
            missing_titles.append(title)

        merged = {
            **(existing_review or {}),
            "id": (existing_review or {}).get("id", f"early-horror-{slug}"),
            "movieTitle": title,
            "slug": slug,
            "releaseYear": year,
            "posterImage": poster_image,
            "backdropImage": (existing_review or {}).get("backdropImage", ""),
            "verdict": verdict_for_rating(rating),
            "rating": rating,
            "reviewerName": normalize_reviewer(choose_row_value(source_row, fix_row, "Reviewer")),
            "quickHit": collapse_spaces(choose_row_value(source_row, fix_row, "Quick Hit")),
            "fullTake": collapse_spaces(choose_row_value(source_row, fix_row, "Full Take")),
            "reviewVideoUrl": (existing_review or {}).get("reviewVideoUrl", ""),
            "whereToWatchUrl": (existing_review or {}).get("whereToWatchUrl", ""),
            "amazonAffiliateUrl": (existing_review or {}).get("amazonAffiliateUrl", ""),
            "createdAt": (existing_review or {}).get("createdAt", now),
            "updatedAt": now,
            "featured": bool((existing_review or {}).get("featured", False)),
            "genreTags": split_tags(choose_row_value(source_row, fix_row, "Genre Tags")),
            "moodTags": split_tags(choose_row_value(source_row, fix_row, "Mood Tags")),
            "runtime": collapse_spaces(choose_row_value(source_row, fix_row, "Runtime")),
            "director": collapse_spaces(choose_row_value(source_row, fix_row, "Director")),
            "status": "published" if poster_image else "draft",
            "collection": COLLECTION,
            "pendingPoster": not bool(poster_image),
        }

        reviews_by_slug[slug] = merged
        if slug not in ordered_imported_slugs:
            ordered_imported_slugs.append(slug)

    preserved_reviews = []
    for review in reviews:
        if review["slug"] not in ordered_imported_slugs:
            preserved_reviews.append(review)

    site_data["reviews"] = preserved_reviews + [
        reviews_by_slug[slug] for slug in ordered_imported_slugs
    ]
    DATA_PATH.write_text(json.dumps(site_data, indent=2, ensure_ascii=False) + "\n")
    build_report(imported_count, live_visible_count, hidden_pending_count, missing_titles)

    print(f"Imported count: {imported_count}")
    print(f"Live visible count: {live_visible_count}")
    print(f"Hidden pending-poster count: {hidden_pending_count}")
    print(f"Still missing after normalization/alias pass: {len(missing_titles)}")
    print(f"Report: {REPORT_PATH}")


if __name__ == "__main__":
    main()
