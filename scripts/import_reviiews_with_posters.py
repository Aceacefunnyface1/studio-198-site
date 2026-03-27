from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path("/Users/bradbehnke/Desktop/movie_posters/reviiews.xlsx")
SOURCE_POSTER_DIR = Path("/Users/bradbehnke/Desktop/movie_posters/posters")
DATA_PATH = Path(__file__).resolve().parents[1] / "data" / "site-data.json"
DEST_POSTER_DIR = Path(__file__).resolve().parents[1] / "public" / "posters" / "reviiews-import"
REPORT_PATH = Path(__file__).resolve().parents[1] / "reports" / "reviiews-import-missing.md"


def normalize_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii").lower()
    ascii_value = ascii_value.replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "", ascii_value)


def without_and(value: str) -> str:
    return value.replace("and", "")


def slugify(value: str) -> str:
    return re.sub(r"(^-+|-+$)", "", re.sub(r"[^a-z0-9]+", "-", value.lower())).strip("-")


def strip_trailing_year(value: str) -> str:
    return re.sub(r"(19|20)\d{2}$", "", value)


def split_tags(value: object) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def parse_rating(value: object) -> float:
    try:
        rating = float(str(value).strip())
    except (TypeError, ValueError):
        return 3.7

    return max(0.0, min(5.0, rating))


def verdict_for_rating(rating: float) -> str:
    if rating >= 4.0:
        return "WATCH"
    if rating >= 3.0:
        return "ONLY IF YOU'RE INTO IT"
    if rating >= 1.1:
        return "DON'T BOTHER"
    return "STRAIGHT TRASH 💩"


def build_poster_lookup() -> dict[str, Path]:
    lookup: dict[str, Path] = {}
    for poster_path in SOURCE_POSTER_DIR.iterdir():
        if not poster_path.is_file():
            continue

        key = normalize_key(poster_path.stem)
        lookup.setdefault(key, poster_path)
    return lookup


def find_matching_poster(title: str, slug: str, poster_lookup: dict[str, Path]) -> Path | None:
    title_key = normalize_key(title)
    slug_key = normalize_key(slug)
    candidates = [
        title_key,
        without_and(title_key),
        strip_trailing_year(title_key),
        without_and(strip_trailing_year(title_key)),
        slug_key,
        without_and(slug_key),
        strip_trailing_year(slug_key),
        without_and(strip_trailing_year(slug_key)),
    ]

    for candidate in candidates:
        if candidate and candidate in poster_lookup:
            return poster_lookup[candidate]

    return None


def read_rows() -> list[dict[str, object]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, object]] = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values or not values[0] or not values[1]:
            continue
        rows.append(dict(zip(headers, values)))

    return rows


def write_missing_report(missing: list[tuple[str, str]]) -> None:
    lines = [
        "# reviiews Import Missing Posters",
        "",
        f"Source workbook: `{WORKBOOK_PATH}`",
        f"Source poster folder: `{SOURCE_POSTER_DIR}`",
        "",
        f"Missing matches: **{len(missing)}**",
        "",
    ]

    if missing:
        for title, slug in missing:
            lines.append(f"- {title} (`{slug}`)")
    else:
        lines.append("- None")

    REPORT_PATH.write_text("\n".join(lines) + "\n")


def main() -> None:
    rows = read_rows()
    poster_lookup = build_poster_lookup()
    site_data = json.loads(DATA_PATH.read_text())
    reviews_by_slug = {review["slug"]: review for review in site_data["reviews"]}
    existing_slugs = set(reviews_by_slug)
    now = datetime.now(timezone.utc).isoformat()

    DEST_POSTER_DIR.mkdir(parents=True, exist_ok=True)

    added = 0
    updated = 0
    missing: list[tuple[str, str]] = []

    for row in rows:
        title = str(row["Title"]).strip()
        slug = str(row["Slug"]).strip() or slugify(title)
        poster_path = find_matching_poster(title, slug, poster_lookup)

        if poster_path is None:
            missing.append((title, slug))
            continue

        destination_name = f"{slug}{poster_path.suffix.lower()}"
        destination_path = DEST_POSTER_DIR / destination_name
        shutil.copy2(poster_path, destination_path)

        rating = parse_rating(row.get("Rating"))
        review = reviews_by_slug.get(slug)
        base = review or {
            "id": f"imported-{slug}",
            "createdAt": now,
            "featured": False,
        }

        merged = {
            **base,
            "movieTitle": title,
            "slug": slug,
            "releaseYear": int(row["Year"]) if str(row.get("Year", "")).strip().isdigit() else None,
            "posterImage": f"/posters/reviiews-import/{destination_name}",
            "backdropImage": base.get("backdropImage", ""),
            "verdict": verdict_for_rating(rating),
            "rating": rating,
            "reviewerName": str(row.get("Reviewer") or "Ace").strip() or "Ace",
            "quickHit": str(row.get("Quick Hit") or "").strip(),
            "fullTake": str(row.get("Full Take") or "").strip(),
            "reviewVideoUrl": str(row.get("Video Review URL") or "").strip(),
            "whereToWatchUrl": str(row.get("Where To Watch URL") or "").strip(),
            "updatedAt": now,
            "featured": base.get("featured", False),
            "genreTags": split_tags(row.get("Genre Tags")),
            "moodTags": split_tags(row.get("Mood Tags")),
            "runtime": str(base.get("runtime", "") or ""),
            "director": str(row.get("Director") or "").strip(),
            "status": "published",
        }

        reviews_by_slug[slug] = merged
        if slug in existing_slugs:
            updated += 1
        else:
            added += 1

    preserved = [
        review for review in site_data["reviews"] if review["slug"] not in reviews_by_slug
    ]
    site_data["reviews"] = [*reviews_by_slug.values(), *preserved]
    DATA_PATH.write_text(json.dumps(site_data, indent=2, ensure_ascii=True) + "\n")
    write_missing_report(missing)

    print(f"Matched and imported: {added + updated}")
    print(f"Added new reviews: {added}")
    print(f"Updated existing reviews: {updated}")
    print(f"Missing poster matches: {len(missing)}")
    print(f"Missing report: {REPORT_PATH}")


if __name__ == "__main__":
    main()
